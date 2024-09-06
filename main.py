import os
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from typing import List, Dict, Optional
from dataclasses import dataclass, field

# Arrays para tipos de datos, constraints y tipos de columna
DATA_TYPES = [
    'INT', 'BIGINT', 'SMALLINT', 'TINYINT', 'BIT',
    'DECIMAL', 'NUMERIC', 'MONEY', 'SMALLMONEY',
    'FLOAT', 'REAL',
    'DATE', 'TIME', 'DATETIME', 'DATETIME2', 'DATETIMEOFFSET', 'SMALLDATETIME',
    'CHAR', 'VARCHAR', 'TEXT',
    'NCHAR', 'NVARCHAR', 'NTEXT',
    'BINARY', 'VARBINARY', 'IMAGE',
    'UNIQUEIDENTIFIER',
    'XML',
    'JSON',
    'ENUM'  # Añadido para soportar ENUM si es necesario
]

CONSTRAINT_TYPES = ['PRIMARY KEY', 'FOREIGN KEY', 'UNIQUE', 'CHECK', 'DEFAULT']

COLUMN_TYPES = ['PK', 'FK', 'UNIQUE']

@dataclass
class ColumnAttribute:
    name: str
    data_type: str
    length: str = ''
    column_type: List[str] = field(default_factory=list)
    default_value: str = ''
    description: str = ''
    constraint_name: str = ''
    constraint_type: str = ''
    constraint_column: str = ''
    reference_table: str = ''
    reference_column: str = ''
    enum_values: List[str] = field(default_factory=list)
    check_constraint: str = ''
    is_nullable: bool = True

@dataclass
class TableObject:
    name: str
    columns: Dict[str, ColumnAttribute] = field(default_factory=dict)

class SQLParser:
    DEFAULT_VALUES = {
        'GETDATE()': 'GETDATE()',
        'CURRENT_TIMESTAMP': 'CURRENT_TIMESTAMP',
        'CURRENT_DATE': 'CURRENT_DATE',
        'NOW()': 'NOW()',
        'SYSDATE()': 'SYSDATE()',
        'LOCALTIMESTAMP': 'LOCALTIMESTAMP',
        'LOCALTIME': 'LOCALTIME',
        'DATE()': 'DATE()',
        'TIME()': 'TIME()',
        'CURDATE()': 'CURDATE()',
        'CURTIME()': 'CURTIME()',
        'UTC_TIMESTAMP()': 'UTC_TIMESTAMP()',
        'UTC_DATE()': 'UTC_DATE()',
        'UTC_TIME()': 'UTC_TIME()',
        'SYSDATETIME()': 'SYSDATETIME()',
        'CURRENT_USER': 'CURRENT_USER',
        'SESSION_USER': 'SESSION_USER',
        'SYSTEM_USER': 'SYSTEM_USER',
        'USER()': 'USER()',
        'UUID()': 'UUID()',
        'RAND()': 'RAND()',
        'MD5()': 'MD5()',
        'SHA1()': 'SHA1()',
        'SHA2()': 'SHA2()',
        'CURRENT_ROLE': 'CURRENT_ROLE',
        'SESSION_USER()': 'SESSION_USER()',
        'SYSTEM_USER()': 'SYSTEM_USER()',
        'DATABASE()': 'DATABASE()',
        'SCHEMA()': 'SCHEMA()',
        'VERSION()': 'VERSION()',
        'NULL': 'NULL'
    }

    @staticmethod
    def clean_value(value: str) -> str:
        return value.strip().strip(',').strip('(').strip(')').strip("'").strip('"')

    @classmethod
    def parse_create_table(cls, sql: str) -> Dict[str, TableObject]:
        tables = {}
        current_table = None
        columns_started = False
        lines = sql.split('\n')
        
        while lines:
            line = lines.pop(0).strip()
            
            if line.lower().startswith('create table'):
                columns_started = True
                match = re.search(r'CREATE\s+TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?(\w+)', line, re.IGNORECASE)
                if match:
                    current_table = match.group(1)
                    tables[current_table] = TableObject(current_table)

            elif line.lower().startswith('constraint'):
                if current_table is None:
                    print("Warning: Found a constraint definition without an active table context.")
                    continue

                cls.parse_constraint(line, tables[current_table])

            elif line and columns_started and not line.startswith(')') and not line.lower().startswith('create'):
                cls.parse_column(line, tables[current_table], lines)

            elif line.startswith(')'):
                columns_started = False
                current_table = None

        return tables

    @classmethod
    def parse_constraint(cls, line: str, table: TableObject):
        constraint_info = re.search(r'CONSTRAINT\s+(\w+)\s+(PRIMARY KEY|FOREIGN KEY|UNIQUE)\s*\((\w+)\)(?:\s+REFERENCES\s+(\w+)\((\w+)\))?', line, re.IGNORECASE)
        if constraint_info:
            constraint_name, constraint_type, column, ref_table, ref_column = constraint_info.groups()
            if column in table.columns:
                table.columns[column].constraint_name = constraint_name
                table.columns[column].constraint_type = constraint_type.upper()
                table.columns[column].constraint_column = column
                table.columns[column].reference_table = ref_table or ''
                table.columns[column].reference_column = ref_column or ''
                if constraint_type.upper() == 'PRIMARY KEY':
                    table.columns[column].column_type.append('PK')
                elif constraint_type.upper() == 'FOREIGN KEY':
                    table.columns[column].column_type.append('FK')
                elif constraint_type.upper() == 'UNIQUE':
                    table.columns[column].column_type.append('UNIQUE')

    @classmethod
    def parse_column(cls, line: str, table: TableObject, lines: list):
        column_info = re.split(r'\s+', line, 2)
        if len(column_info) >= 2:
            name = cls.clean_value(column_info[0])
            data_type_info = ' '.join(column_info[1:])  # Join all remaining parts
            
            enum_values, data_type, length = cls.parse_data_type(data_type_info)

            column_type = []
            default_value = cls.parse_default_value(line)
            check_constraint = cls.parse_check_constraint(line)
            is_nullable = 'NOT NULL' not in line.upper()

            if 'PRIMARY KEY' in line.upper():
                column_type.append('PK')
            if 'UNIQUE' in line.upper():
                column_type.append('UNIQUE')
            if check_constraint:
                column_type.append('CHECK')

            description = cls.parse_description(line)

            column = ColumnAttribute(
                name=name,
                data_type=data_type,
                length=length,
                column_type=column_type,
                default_value=default_value,
                description=description,
                enum_values=enum_values,
                check_constraint=check_constraint,
                is_nullable=is_nullable
            )
            table.columns[name] = column

    @classmethod
    def parse_data_type(cls, data_type_info: str) -> tuple:
        enum_values = []
        # Improved ENUM detection
        enum_match = re.search(r'ENUM\s*\((.*?)\)(?:\s+DEFAULT\s+.*?)?(?:,|$)', data_type_info, re.IGNORECASE | re.DOTALL)
        if enum_match:
            # Handle multiple lines and whitespace
            enum_str = enum_match.group(1)
            enum_values = [cls.clean_value(v) for v in re.findall(r"'([^']*)'", enum_str)]
            data_type = 'ENUM'
            length = ''
        else:
            # Handle other data types
            parts = data_type_info.split('(', 1)
            data_type = parts[0].strip().upper()
            if len(parts) > 1:
                length_part = parts[1].split(')', 1)[0]
                length = length_part.strip()
            else:
                length = ''

        return enum_values, data_type, length

    @classmethod
    def parse_default_value(cls, line: str) -> str:
        default_match = re.search(r'DEFAULT\s+(.+?)(?:\s+|$)', line, re.IGNORECASE)
        if default_match:
            found_default = cls.clean_value(default_match.group(1))
            return cls.DEFAULT_VALUES.get(found_default.upper(), found_default)
        return ''

    @classmethod
    def parse_check_constraint(cls, line: str) -> str:
        check_match = re.search(r'CHECK\s*\((.*?)\)', line, re.IGNORECASE)
        return check_match.group(1).strip() if check_match else ''

    @classmethod
    def parse_description(cls, line: str) -> str:
        comment_match = re.search(r'--\s*(.+)$', line)
        return comment_match.group(1).strip() if comment_match else ''

class ExcelGenerator:
    @staticmethod
    def create_excel(tables: Dict[str, TableObject], output_file: str):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        headers = ['Variable', 'Tipo', 'Longitud', 'Tipo de Columna', 'Default', 'Descripción', 
                   'Constraint Name', 'Tipo de Constraint', 'Columna', 'Referencia Tabla', 'Referencia Columna', 'Valores ENUM', 'Check Constraint', 'NOT NULL']

        for table_name, table in tables.items():
            ws = wb.create_sheet(title=table_name[:31])
            
            ExcelGenerator.set_headers(ws, headers)
            ExcelGenerator.fill_data(ws, table)
            ExcelGenerator.format_worksheet(ws)

        wb.save(output_file)
        os.startfile(output_file)

    @staticmethod
    def set_headers(ws, headers):
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    @staticmethod
    def fill_data(ws, table):
        for row, (column_name, column) in enumerate(table.columns.items(), start=2):
            ws.cell(row=row, column=1, value=SQLParser.clean_value(column.name)).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=2, value=SQLParser.clean_value(column.data_type).split()[0]).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=SQLParser.clean_value(column.length)).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=4, value=SQLParser.clean_value(', '.join(column.column_type))).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=SQLParser.clean_value(column.default_value)).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=6, value=SQLParser.clean_value(column.description)).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=7, value=SQLParser.clean_value(column.constraint_name)).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=8, value=SQLParser.clean_value(column.constraint_type)).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=9, value=SQLParser.clean_value(column.constraint_column)).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=10, value=SQLParser.clean_value(column.reference_table)).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=11, value=SQLParser.clean_value(column.reference_column)).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=12, value=SQLParser.clean_value(', '.join(column.enum_values))).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=13, value=SQLParser.clean_value(column.check_constraint)).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=14, value=str(not column.is_nullable)).alignment = Alignment(horizontal='center')

    @staticmethod
    def format_worksheet(ws):
        for col in range(ws.max_column, 0, -1):
            if all(ws.cell(row=row, column=col).value in (None, '') for row in range(1, ws.max_row + 1)):
                ws.delete_cols(col)

        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = max_length + 2

def main():
    try:
        with open('data.sql', 'r', encoding='utf-8') as file:
            sql_content = file.read()

        tables = SQLParser.parse_create_table(sql_content)
        ExcelGenerator.create_excel(tables, 'diccionario_datos_avanzado.xlsx')
        print("El diccionario de datos ha sido generado en 'diccionario_datos_avanzado.xlsx'")
    except Exception as e:
        print(f"Se produjo un error: {str(e)}")

if __name__ == "__main__":
    main()