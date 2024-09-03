import os
import re
from sql_data_types import parse_column_data_type
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from typing import List, Dict, Any, Optional

class Column:
    def __init__(self, name: str, data_type: str, length: str, col_type: List[str], 
                 default_value: str = '', description: str = '', enum_values: List[str] = None,
                 check_constraint: str = ''):
        self.name = name
        self.data_type = data_type
        self.length = length
        self.col_type = col_type
        self.default_value = default_value
        self.description = description
        self.enum_values = enum_values or []
        self.check_constraint = check_constraint
        self.constraint_name = ''
        self.constraint_type = ''
        self.constraint_column = ''
        self.reference_table = ''
        self.reference_column = ''

class Constraint:
    def __init__(self, name: str, constraint_type: str, column: str, 
                 reference_table: str = '', reference_column: str = ''):
        self.name = name
        self.constraint_type = constraint_type
        self.column = column
        self.reference_table = reference_table
        self.reference_column = reference_column

class Table:
    def __init__(self, name: str):
        self.name = name
        self.columns: List[Column] = []
        self.constraints: List[Constraint] = []

    def add_column(self, column: Column):
        self.columns.append(column)

    def add_constraint(self, constraint: Constraint):
        self.constraints.append(constraint)
        for column in self.columns:
            if column.name == constraint.column:
                column.constraint_name = constraint.name
                column.constraint_type = constraint.constraint_type
                column.constraint_column = constraint.column
                column.reference_table = constraint.reference_table
                column.reference_column = constraint.reference_column

def clean_value(value: str) -> str:
    return value.strip().strip(',').strip('(').strip(')').strip("'").strip('"')

def parse_create_table(sql: str) -> Dict[str, Table]:
    tables = {}
    current_table = None
    for line in sql.split('\n'):
        line = line.strip()
        if line.lower().startswith('create table'):
            match = re.search(r'CREATE\s+TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?(\w+)', line, re.IGNORECASE)
            if match:
                current_table = match.group(1)
                tables[current_table] = Table(current_table)
        elif line and current_table and not line.startswith(')') and not line.startswith('--'):
            if line.lower().startswith('constraint'):
                constraint_info = re.search(r'CONSTRAINT\s+(\w+)\s+(PRIMARY KEY|FOREIGN KEY|UNIQUE)\s*\((\w+)\)(?:\s+REFERENCES\s+(\w+)\((\w+)\))?', line, re.IGNORECASE)
                if constraint_info:
                    constraint_name, constraint_type, column, ref_table, ref_column = constraint_info.groups()
                    constraint = Constraint(
                        name=constraint_name,
                        constraint_type=constraint_type.upper(),
                        column=column,
                        reference_table=ref_table,
                        reference_column=ref_column
                    )
                    tables[current_table].add_constraint(constraint)
            else:
                column_info = re.split(r'\s+', line, 2)
                if len(column_info) >= 2:
                    name = clean_value(column_info[0])
                    data_type_info = column_info[1].upper()
                    
                    enum_values = []
                    enum_match = re.search(r'ENUM\((.*?)\)', data_type_info, re.IGNORECASE)
                    if enum_match:
                        enum_values = [clean_value(v) for v in enum_match.group(1).split(',')]
                        data_type = 'ENUM'
                        length = ''
                    else:
                        data_type = data_type_info.split('(')[0]
                        length_match = re.search(r'\((\d+(?:,\s*\d+)?)\)', data_type_info)
                        length = length_match.group(1) if length_match else ''
                    
                    col_type = []
                    if 'PRIMARY KEY' in line.upper():
                        col_type.append('PK')
                    if 'UNIQUE' in line.upper():
                        col_type.append('UNIQUE')
                    if 'NOT NULL' in line.upper():
                        col_type.append('NOT NULL')
                    
                    default_match = re.search(r'DEFAULT\s+(.+?)(?:\s+|$)', line, re.IGNORECASE)
                    default_value = clean_value(default_match.group(1)) if default_match else ''
                    
                    check_constraint = ''
                    check_match = re.search(r'CHECK\s*\((.*?)\)', line, re.IGNORECASE)
                    if check_match:
                        check_constraint = check_match.group(1).strip()
                        col_type.append('CHECK')
                    
                    description = ''
                    comment_match = re.search(r'--\s*(.+)$', line)
                    if comment_match:
                        description = comment_match.group(1).strip()
                    
                    column = Column(
                        name=name,
                        data_type=data_type,
                        length=length,
                        col_type=col_type,
                        default_value=default_value,
                        description=description,
                        enum_values=enum_values,
                        check_constraint=check_constraint
                    )
                    tables[current_table].add_column(column)
    return tables

def create_excel(tables: Dict[str, Table]):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    headers = ['Variable', 'Tipo', 'Longitud', 'Tipo de Columna', 'Default', 'Descripción', 
               'Constraint Name', 'Tipo de Constraint', 'Columna', 'Referencia Tabla', 'Referencia Columna', 'Valores ENUM', 'Check Constraint']

    for table_name, table in tables.items():
        ws = wb.create_sheet(title=table_name[:31])
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        for row, column in enumerate(table.columns, start=2):
            ws.cell(row=row, column=1, value=column.name).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=2, value=column.data_type).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=column.length).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=4, value=', '.join(column.col_type)).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=column.default_value).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=6, value=column.description).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=7, value=column.constraint_name).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=8, value=column.constraint_type).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=9, value=column.constraint_column).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=10, value=column.reference_table).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=11, value=column.reference_column).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=12, value=', '.join(column.enum_values)).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=13, value=column.check_constraint).alignment = Alignment(horizontal='center')

        # Eliminar columnas vacías
        for col in range(ws.max_column, 0, -1):
            if all(ws.cell(row=row, column=col).value in (None, '') for row in range(1, ws.max_row + 1)):
                ws.delete_cols(col)

        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = max_length + 2

    # Guardar el archivo Excel
    excel_file_path = 'diccionario_datos_mejorado.xlsx'
    wb.save(excel_file_path)

    # Abrir el archivo Excel automáticamente
    os.startfile(excel_file_path)

# Leer el contenido del archivo SQL
with open('data.sql', 'r') as file:
    sql_content = file.read()

# Parsear las tablas
tables = parse_create_table(sql_content)

# Crear el archivo Excel
create_excel(tables)

print("El diccionario de datos ha sido generado en 'diccionario_datos_mejorado.xlsx'")